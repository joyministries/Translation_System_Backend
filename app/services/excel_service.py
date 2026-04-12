import os
from typing import Any

import openpyxl
import xlrd

from app.config import settings


def parse_excel(file_path: str) -> dict[str, Any]:
    full_path = os.path.join(settings.STORAGE_ROOT, file_path)

    if file_path.endswith(".xls"):
        try:
            return parse_xls(full_path)
        except Exception:
            return {"sheet_names": ["Sheet1"], "sheets": {}, "total_sheets": 1}

    wb = openpyxl.load_workbook(full_path, data_only=True)

    sheets_data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(cell) if cell is not None else "" for cell in row])
        sheets_data[sheet_name] = rows

    wb.close()

    return {
        "sheet_names": list(sheets_data.keys()),
        "sheets": sheets_data,
        "total_sheets": len(sheets_data),
    }


def parse_xls(file_path: str) -> dict[str, Any]:
    wb = xlrd.open_workbook(file_path)

    sheets_data = {}
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for row_idx in range(ws.nrows):
            row = []
            for col_idx in range(ws.ncols):
                cell = ws.cell_value(row_idx, col_idx)
                row.append(str(cell) if cell else "")
            if any(row):
                rows.append(row)
        sheets_data[sheet_name] = rows

    return {
        "sheet_names": list(wb.sheet_names()),
        "sheets": sheets_data,
        "total_sheets": len(wb.sheet_names()),
    }
